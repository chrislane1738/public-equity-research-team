from pathlib import Path

from openpyxl import load_workbook

from backend.tools.xlsx_writer import write_dcf_xlsx, write_comps_xlsx


def test_write_dcf_xlsx_creates_all_required_tabs(tmp_path):
    out = tmp_path / "dcf.xlsx"
    write_dcf_xlsx(
        path=out,
        ticker="NVDA",
        wacc=10.5,
        revenue_build=[
            {"year": 2026, "revenue": 80_000, "growth_pct": 25.0,
             "segments": {"data_center": 60_000, "gaming": 12_000, "pro_viz": 4_000, "auto": 4_000}},
            {"year": 2027, "revenue": 90_000, "growth_pct": 12.5,
             "segments": {"data_center": 70_000, "gaming": 12_000, "pro_viz": 4_000, "auto": 4_000}},
        ],
        op_model=[
            {"year": 2026, "gross_margin_pct": 73, "rd_pct": 18, "sm_pct": 8, "ga_pct": 3,
             "ebit": 32_000, "ebit_margin_pct": 40},
            {"year": 2027, "gross_margin_pct": 73, "rd_pct": 18, "sm_pct": 8, "ga_pct": 3,
             "ebit": 36_000, "ebit_margin_pct": 40},
        ],
        fcf=[
            {"year": 2026, "ebit": 32_000, "nopat": 25_280, "da": 4_000, "capex": 5_600,
             "wc_change": 800, "fcf": 22_880},
            {"year": 2027, "ebit": 36_000, "nopat": 28_440, "da": 4_500, "capex": 6_300,
             "wc_change": 900, "fcf": 25_740},
        ],
        wacc_inputs={"beta": 1.6, "rf": 4.25, "erp": 5.5, "cost_of_debt": 5.0,
                     "tax_rate": 0.21, "weight_equity": 0.95, "weight_debt": 0.05,
                     "wacc": 10.5},
        ggm={"growth": 2.5, "fcf_t": 25_740, "tv": 350_000, "pv_tv": 200_000,
             "ev": 300_000, "equity": 290_000, "implied_price": 116.0},
        exit_mult={"peer_median_multiple": 22.0, "haircut": 0.85, "applied_multiple": 18.7,
                   "ebitda_t": 38_000, "tv": 710_600, "pv_tv": 410_000, "ev": 510_000,
                   "equity": 500_000, "implied_price": 200.0},
        blend={"weight_ggm": 0.5, "ggm_implied_price": 116.0, "exit_implied_price": 200.0,
               "blended_price": 158.0},
        sensitivity_ggm={(9, 2): 100, (10, 2): 95, (11, 2): 90,
                         (9, 3): 110, (10, 3): 105, (11, 3): 100},
        sensitivity_exit={(9, 17): 150, (10, 17): 140, (11, 17): 130,
                          (9, 20): 175, (10, 20): 165, (11, 20): 155},
        summary={"rating": "Buy", "blended_pt": 158.0, "current_price": 110.0,
                 "upside_pct": 43.6},
    )
    assert out.exists()
    wb = load_workbook(out)
    expected_tabs = ["Cover", "Revenue Build", "Operating Model", "FCF", "WACC",
                     "DCF — GGM", "DCF — Exit Mult", "DCF — Blend",
                     "Sensitivities", "Summary"]
    for tab in expected_tabs:
        assert tab in wb.sheetnames, f"missing tab: {tab}"


def test_write_dcf_xlsx_summary_tab_contains_blended_pt(tmp_path):
    out = tmp_path / "dcf.xlsx"
    write_dcf_xlsx(
        path=out, ticker="NVDA", wacc=10.0,
        revenue_build=[], op_model=[], fcf=[],
        wacc_inputs={"beta": 1.0, "rf": 4.0, "erp": 5.5, "cost_of_debt": 5.0,
                     "tax_rate": 0.21, "weight_equity": 1.0, "weight_debt": 0.0,
                     "wacc": 10.0},
        ggm={"growth": 2, "fcf_t": 100, "tv": 1000, "pv_tv": 500,
             "ev": 1000, "equity": 1000, "implied_price": 100.0},
        exit_mult={"peer_median_multiple": 20, "haircut": 0.85, "applied_multiple": 17,
                   "ebitda_t": 100, "tv": 1700, "pv_tv": 800, "ev": 1500,
                   "equity": 1500, "implied_price": 150.0},
        blend={"weight_ggm": 0.5, "ggm_implied_price": 100.0,
               "exit_implied_price": 150.0, "blended_price": 125.0},
        sensitivity_ggm={}, sensitivity_exit={},
        summary={"rating": "Hold", "blended_pt": 125.0,
                 "current_price": 110.0, "upside_pct": 13.6},
    )
    wb = load_workbook(out)
    cells = [c.value for c in wb["Summary"]["A"]]
    assert any(v == "Blended PT" for v in cells)
    # Confirm the value cell next to "Blended PT" actually holds the input.
    sm = wb["Summary"]
    label_row = next(i for i, row in enumerate(sm.iter_rows(values_only=True), start=1)
                     if row and row[0] == "Blended PT")
    assert sm.cell(row=label_row, column=2).value == 125.0


def test_write_comps_xlsx_creates_required_tabs(tmp_path):
    out = tmp_path / "comps.xlsx"
    write_comps_xlsx(
        path=out, ticker="NVDA",
        peers=[
            {"symbol": "NVDA", "market_cap": 3e12, "ev_to_ebitda": 45, "pe": 80, "ev_to_sales": 22},
            {"symbol": "AMD", "market_cap": 250e9, "ev_to_ebitda": 30, "pe": 50, "ev_to_sales": 8},
            {"symbol": "INTC", "market_cap": 150e9, "ev_to_ebitda": 12, "pe": 18, "ev_to_sales": 3},
        ],
        summary={
            "ev_to_ebitda": {"median": 30, "p25": 21, "p75": 37.5, "n": 3},
            "pe":         {"median": 50, "p25": 34, "p75": 65, "n": 3},
            "ev_to_sales":{"median": 8,  "p25": 5.5,"p75": 15, "n": 3},
        },
    )
    wb = load_workbook(out)
    for tab in ["Cover", "Peers", "Summary"]:
        assert tab in wb.sheetnames
