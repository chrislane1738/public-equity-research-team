"""openpyxl-based xlsx writer for DCF and Comps workbooks.

Each worksheet is written explicitly — no template files required. Plan B keeps
formatting minimal (bold headers, USD/percent number formats); a future task can
add color/conditional-format polish.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


HEADER = Font(bold=True)


def _ensure_dir(path: Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


def _write_kv_block(ws, start_row: int, kvs: list[tuple[str, object]],
                    label_col: str = "A", value_col: str = "B") -> int:
    r = start_row
    for k, v in kvs:
        ws[f"{label_col}{r}"] = k
        ws[f"{label_col}{r}"].font = HEADER
        ws[f"{value_col}{r}"] = v
        r += 1
    return r


def _table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row, column=j + 1, value=h)
        c.font = HEADER
    for i, row in enumerate(rows, start=1):
        for j, v in enumerate(row):
            ws.cell(row=start_row + i, column=j + 1, value=v)
    return start_row + len(rows) + 1


def write_dcf_xlsx(
    path: Path,
    ticker: str,
    wacc: float,
    revenue_build: list[dict],
    op_model: list[dict],
    fcf: list[dict],
    wacc_inputs: dict,
    ggm: dict,
    exit_mult: dict,
    blend: dict,
    sensitivity_ggm: dict[tuple[float, float], float],
    sensitivity_exit: dict[tuple[float, float], float],
    summary: dict,
) -> None:
    _ensure_dir(path)
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover")
    _write_kv_block(cover, 1, [
        ("Ticker", ticker),
        ("Model", "DCF (GGM + Exit Mult + Blend)"),
        ("WACC", wacc),
    ])

    rb = wb.create_sheet("Revenue Build")
    headers = ["Year", "Revenue", "Growth %"]
    seg_keys: list[str] = []
    if revenue_build:
        seg_keys = sorted({k for r in revenue_build for k in r.get("segments", {}).keys()})
    headers.extend(seg_keys)
    rows = [[r["year"], r["revenue"], r["growth_pct"]] +
            [r.get("segments", {}).get(k, "") for k in seg_keys]
            for r in revenue_build]
    _table(rb, 1, headers, rows)

    om = wb.create_sheet("Operating Model")
    _table(om, 1,
           ["Year", "Gross margin %", "R&D %", "S&M %", "G&A %", "EBIT", "EBIT margin %"],
           [[r["year"], r["gross_margin_pct"], r["rd_pct"], r["sm_pct"], r["ga_pct"],
             r["ebit"], r["ebit_margin_pct"]] for r in op_model])

    fc = wb.create_sheet("FCF")
    _table(fc, 1,
           ["Year", "EBIT", "NOPAT", "D&A", "Capex", "ΔWC", "FCF"],
           [[r["year"], r["ebit"], r["nopat"], r["da"], r["capex"],
             r["wc_change"], r["fcf"]] for r in fcf])

    wsh = wb.create_sheet("WACC")
    _write_kv_block(wsh, 1, [
        ("Beta", wacc_inputs["beta"]),
        ("Rf (10Y UST, %)", wacc_inputs["rf"]),
        ("ERP (%)", wacc_inputs["erp"]),
        ("Pre-tax cost of debt (%)", wacc_inputs["cost_of_debt"]),
        ("Tax rate", wacc_inputs["tax_rate"]),
        ("Weight equity", wacc_inputs["weight_equity"]),
        ("Weight debt", wacc_inputs["weight_debt"]),
        ("WACC (%)", wacc_inputs["wacc"]),
    ])

    ggm_ws = wb.create_sheet("DCF — GGM")
    _write_kv_block(ggm_ws, 1, [
        ("Terminal growth (%)", ggm["growth"]),
        ("FCF_T", ggm["fcf_t"]),
        ("Terminal value", ggm["tv"]),
        ("PV of TV", ggm["pv_tv"]),
        ("EV", ggm["ev"]),
        ("Equity", ggm["equity"]),
        ("Implied price", ggm["implied_price"]),
    ])

    exit_ws = wb.create_sheet("DCF — Exit Mult")
    _write_kv_block(exit_ws, 1, [
        ("Peer median EV/EBITDA", exit_mult["peer_median_multiple"]),
        ("Haircut", exit_mult["haircut"]),
        ("Applied multiple", exit_mult["applied_multiple"]),
        ("EBITDA_T", exit_mult["ebitda_t"]),
        ("Terminal value", exit_mult["tv"]),
        ("PV of TV", exit_mult["pv_tv"]),
        ("EV", exit_mult["ev"]),
        ("Equity", exit_mult["equity"]),
        ("Implied price", exit_mult["implied_price"]),
    ])

    blend_ws = wb.create_sheet("DCF — Blend")
    _write_kv_block(blend_ws, 1, [
        ("Weight on GGM", blend["weight_ggm"]),
        ("GGM implied price", blend["ggm_implied_price"]),
        ("Exit implied price", blend["exit_implied_price"]),
        ("Blended price", blend["blended_price"]),
    ])

    sens = wb.create_sheet("Sensitivities")
    sens["A1"] = "GGM: rows = WACC, cols = terminal g"
    sens["A1"].font = HEADER
    ggm_ys = sorted({k[0] for k in sensitivity_ggm.keys()}) if sensitivity_ggm else []
    if sensitivity_ggm:
        ggm_xs = sorted({k[1] for k in sensitivity_ggm.keys()})
        for j, x in enumerate(ggm_xs):
            sens.cell(row=2, column=j + 2, value=x)
        for i, y in enumerate(ggm_ys):
            sens.cell(row=3 + i, column=1, value=y)
            for j, x in enumerate(ggm_xs):
                sens.cell(row=3 + i, column=j + 2,
                          value=sensitivity_ggm.get((y, x), ""))
    base_row = 3 + len(ggm_ys) + 2
    sens.cell(row=base_row, column=1, value="Exit: rows = WACC, cols = exit multiple").font = HEADER
    if sensitivity_exit:
        ex_ys = sorted({k[0] for k in sensitivity_exit.keys()})
        ex_xs = sorted({k[1] for k in sensitivity_exit.keys()})
        for j, x in enumerate(ex_xs):
            sens.cell(row=base_row + 1, column=j + 2, value=x)
        for i, y in enumerate(ex_ys):
            sens.cell(row=base_row + 2 + i, column=1, value=y)
            for j, x in enumerate(ex_xs):
                sens.cell(row=base_row + 2 + i, column=j + 2,
                          value=sensitivity_exit.get((y, x), ""))

    sm = wb.create_sheet("Summary")
    _write_kv_block(sm, 1, [
        ("Rating", summary["rating"]),
        ("Blended PT", summary["blended_pt"]),
        ("Current price", summary["current_price"]),
        ("Upside %", summary["upside_pct"]),
    ])

    wb.save(path)


def write_comps_xlsx(
    path: Path,
    ticker: str,
    peers: list[dict],
    summary: dict[str, dict[str, float]],
) -> None:
    _ensure_dir(path)
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover")
    _write_kv_block(cover, 1, [
        ("Ticker", ticker),
        ("Sheet purpose", "Comparable company analysis — peer multiples"),
    ])

    pe = wb.create_sheet("Peers")
    headers = ["Symbol", "Market cap", "EV/EBITDA", "P/E", "EV/Sales"]
    rows = [[p["symbol"], p["market_cap"], p.get("ev_to_ebitda"),
             p.get("pe"), p.get("ev_to_sales")] for p in peers]
    _table(pe, 1, headers, rows)

    sm = wb.create_sheet("Summary")
    sm["A1"] = "Multiple"; sm["A1"].font = HEADER
    sm["B1"] = "Median"; sm["B1"].font = HEADER
    sm["C1"] = "P25"; sm["C1"].font = HEADER
    sm["D1"] = "P75"; sm["D1"].font = HEADER
    sm["E1"] = "n"; sm["E1"].font = HEADER
    r = 2
    for metric, stats in summary.items():
        sm.cell(row=r, column=1, value=metric)
        sm.cell(row=r, column=2, value=stats.get("median"))
        sm.cell(row=r, column=3, value=stats.get("p25"))
        sm.cell(row=r, column=4, value=stats.get("p75"))
        sm.cell(row=r, column=5, value=stats.get("n"))
        r += 1

    wb.save(path)
